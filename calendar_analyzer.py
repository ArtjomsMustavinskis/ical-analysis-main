import os
import argparse
from datetime import datetime, timezone, timedelta
import re
from icalendar import Calendar
from dateutil import tz
from typing import Dict, List, Pattern, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter


class CalendarAnalyzer:
    def __init__(self, calendar_files: List[str]):
        self.calendar_files = [os.path.join(os.getcwd(), f) for f in calendar_files]
        self.local_tz = tz.gettz('America/Los_Angeles')
        self._calendars = None

    @property
    def calendars(self) -> List[Calendar]:
        if self._calendars is None:
            self._calendars = [self.load_calendar(path) for path in self.calendar_files]
        return self._calendars

    def load_calendar(self, path: str) -> Calendar:
        with open(path, 'rb') as f:
            return Calendar.from_ical(f.read())

    def analyze_events(
        self,
        start_time: datetime,
        end_time: datetime,
        patterns: Dict[str, Pattern]
    ) -> Dict[str, List[Tuple[datetime, str, timedelta]]]:
        events_data = {pattern_name: [] for pattern_name in patterns}
        unmatched_events = []

        for cal in self.calendars:
            for component in cal.walk():
                if component.name != "VEVENT":
                    continue

                dtstart = component.get('dtstart')
                if not dtstart:
                    continue
                event_start = dtstart.dt

                dtend = component.get('dtend')
                if not dtend:
                    event_end = event_start + timedelta(hours=1)
                else:
                    event_end = dtend.dt

                is_all_day = not isinstance(event_start, datetime)
                if isinstance(event_start, datetime):
                    event_start = event_start.replace(tzinfo=timezone.utc).astimezone(self.local_tz)
                else:
                    event_start = datetime.combine(event_start, datetime.min.time(), tzinfo=self.local_tz)

                if isinstance(event_end, datetime):
                    event_end = event_end.replace(tzinfo=timezone.utc).astimezone(self.local_tz)
                else:
                    event_end = datetime.combine(event_end, datetime.min.time(), tzinfo=self.local_tz)

                if event_end < start_time or event_start > end_time:
                    continue

                duration = timedelta(0) if is_all_day else (event_end - event_start)

                summary = str(component.get('summary', ''))
                description = str(component.get('description', ''))
                location = str(component.get('location', ''))
                searchable_text = f"{summary} {description} {location}"

                matched = False
                for pattern_name, regex in patterns.items():
                    if regex.search(searchable_text):
                        events_data[pattern_name].append((event_start, summary, duration))
                        matched = True

                if not matched:
                    unmatched_events.append((event_start, summary, duration))

        if unmatched_events:
            print("\nEvents that did not fit the patterns:")
            for start, summ, dur in unmatched_events:
                print(f"{start.strftime('%Y-%m-%d %H:%M')} | {summ} | {dur}")

        return events_data

    def get_day_stats(self, events_data):
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        distribution = {
            pattern: {
                day: {'count': 0, 'total_hours': 0.0, 'avg_hours': 0.0}
                for day in days
            }
            for pattern in events_data
        }
        for pattern, events in events_data.items():
            for event_start, _, duration in events:
                day = days[event_start.weekday()]
                hours = duration.total_seconds() / 3600
                distribution[pattern][day]['count'] += 1
                distribution[pattern][day]['total_hours'] += hours

        for pattern in distribution:
            for day in days:
                count = distribution[pattern][day]['count']
                if count > 0:
                    distribution[pattern][day]['avg_hours'] = (
                        distribution[pattern][day]['total_hours'] / count
                    )

        return distribution

    def get_time_spent(self, events_data):
        return {
            pattern: sum((duration for _, _, duration in events), timedelta())
            for pattern, events in events_data.items()
        }

    def get_weekly_stats(self, events_data):
        weekly_stats = {}
        for pattern, events in events_data.items():
            weekly_stats[pattern] = {}
            for event_start, _, duration in events:
                monday = event_start - timedelta(days=event_start.weekday())
                week_key = monday.strftime('%Y-%m-%d')
                if week_key not in weekly_stats[pattern]:
                    weekly_stats[pattern][week_key] = {'total_hours': 0.0, 'avg_hours': 0.0}
                weekly_stats[pattern][week_key]['total_hours'] += duration.total_seconds() / 3600

            for week in weekly_stats[pattern].values():
                week['avg_hours'] = week['total_hours'] / 7
        return weekly_stats

    def get_monthly_stats(self, events_data):
        monthly_stats = {}
        for pattern, events in events_data.items():
            monthly_stats[pattern] = {}
            for event_start, _, duration in events:
                month_key = event_start.strftime('%Y-%m')
                if month_key not in monthly_stats[pattern]:
                    monthly_stats[pattern][month_key] = {'total_hours': 0.0, 'avg_hours': 0.0, 'event_count': 0}
                monthly_stats[pattern][month_key]['total_hours'] += duration.total_seconds() / 3600
                monthly_stats[pattern][month_key]['event_count'] += 1

            for month_key, stats in monthly_stats[pattern].items():
                year, month = map(int, month_key.split('-'))
                if month == 12:
                    next_month = datetime(year + 1, 1, 1)
                else:
                    next_month = datetime(year, month + 1, 1)
                days_in_month = (next_month - datetime(year, month, 1)).days
                weeks_in_month = days_in_month / 7.0
                stats['avg_hours'] = stats['total_hours'] / weeks_in_month
        return monthly_stats


def load_or_generate_patterns(ics_files, patterns_file="patterns.txt"):
    if os.path.exists(patterns_file):
        print(f"Loading patterns from {patterns_file}...")
        patterns = {}
        with open(patterns_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or ":" not in line:
                    continue
                name, regex = line.split(":", 1)
                patterns[name.strip()] = re.compile(regex.strip(), re.IGNORECASE)
        return patterns

    print(f"{patterns_file} not found. Generated automatically...")
    keywords = set()

    for file_path in ics_files:
        with open(file_path, "rb") as f:
            cal = Calendar.from_ical(f.read())
        for component in cal.walk():
            if component.name != "VEVENT":
                continue
            text_parts = [
                str(component.get("summary", "")),
                str(component.get("description", "")),
                str(component.get("location", "")),
            ]
            text = " ".join(text_parts)
            words = re.findall(r"[A-Za-zА-Яа-я0-9]{3,}", text)
            for w in words:
                keywords.add(w.lower())

    with open(patterns_file, "w", encoding="utf-8") as f:
        for word in sorted(keywords):
            f.write(f"{word}:(?i){word}\n")

    print(f"Generated {patterns_file}. Edit it and run the script again.")
    return {}


def parse_args():
    parser = argparse.ArgumentParser(description="Analyze multiple .ics calendar files")
    parser.add_argument("files", nargs="+", help="Paths to .ics calendar files")
    parser.add_argument("--start", required=True, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end", required=True, help="End date (YYYY-MM-DD)")
    parser.add_argument("--output", default="calendar_analysis.xlsx", help="Output Excel file")
    return parser.parse_args()


if __name__ == '__main__':
    args = parse_args()

    start_time = datetime.strptime(args.start, "%Y-%m-%d").replace(tzinfo=tz.gettz('America/Los_Angeles'))
    end_time = datetime.strptime(args.end, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=tz.gettz('America/Los_Angeles'))

    patterns = load_or_generate_patterns(args.files)
    if not patterns:
        exit(0)

    analyzer = CalendarAnalyzer(args.files)
    results = analyzer.analyze_events(start_time, end_time, patterns)

    day_dist = analyzer.get_day_stats(results)
    time_spent = analyzer.get_time_spent(results)
    weekly_stats = analyzer.get_weekly_stats(results)
    monthly_stats = analyzer.get_monthly_stats(results)

    print("\nEvent Distribution by Day:")
    for pattern, dist in day_dist.items():
        print(f"\n{pattern}:")
        for day, stats in dist.items():
            print(f"  {day}: {stats['count']} events, {stats['total_hours']:.1f} hours, {stats['avg_hours']:.1f} hours/event")

    print("\nTime Spent on Each Event Type:")
    for pattern, duration in time_spent.items():
        hours = duration.total_seconds() / 3600
        print(f"{pattern}: {hours:.1f} hours")

    print("\nWeekly Statistics:")
    for pattern, stats in weekly_stats.items():
        print(f"\n{pattern}:")
        for week, week_stats in stats.items():
            print(f"  Week of {week}: {week_stats['total_hours']:.1f} hours, {week_stats['avg_hours']:.1f} hours/day")

    print("\nMonthly Statistics:")
    for pattern, stats in monthly_stats.items():
        print(f"\n{pattern}:")
        for month, month_stats in stats.items():
            print(f"  Month of {month}: {month_stats['total_hours']:.1f} hours, {month_stats['avg_hours']:.1f} hours/week, {month_stats['event_count']} events")

    date_list = [start_time + timedelta(days=i) for i in range((end_time - start_time).days + 1)]
    date_str_list = [f"{d.day}-{d.month}-{d.year}" for d in date_list]

    df_data = {}
    for pattern in patterns.keys():
        hours_by_day = []
        for day in date_list:
            total_hours = sum(
                (duration.total_seconds() / 3600)
                for event_start, _, duration in results[pattern]
                if event_start.date() == day.date()
            )
            hours_by_day.append(round(total_hours, 1) if total_hours > 0 else None)
        df_data[pattern] = hours_by_day

    df = pd.DataFrame(df_data, index=date_str_list).T
    df.index.name = "Date / Regex pattern"
    df = df.fillna("-")

    excel_path = args.output
    df.to_excel(excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active

    fill_blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_green = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_red = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for col in range(2, ws.max_column + 1):
        ws.cell(row=1, column=col).fill = fill_green

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).fill = fill_blue

    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill_red
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "B2"

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    wb.save(excel_path)
    print(f"\nExcel file saved as {excel_path}")
